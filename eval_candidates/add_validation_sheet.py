"""
add_validation_sheet.py
Queries processed.meddra_terms for all PT candidates and adds
a MedDRA_Validation sheet + fills meddra_pt_code column in the main sheet.
Run on Hetzner: python3 add_validation_sheet.py
"""

import psycopg2
import openpyxl
from openpyxl.styles import PatternFill, Font
import os

DB_URL = os.environ.get("DATABASE_URL", "postgresql://vigilex:vigilex@localhost:5432/vigilex")

PT_CANDIDATES = [
    "Hyperglycaemia",
    "Diabetic ketoacidosis",
    "Hypoglycaemia",
    "Loss of consciousness",
    "Hypoglycaemic unconsciousness",
    "Nausea",
    "Vomiting",
    "Headache",
    "Delirium",
    "Lethargy",
    "Dizziness",
    "Hyperhidrosis",
    "Skin irritation",
    "Infusion site swelling",
    "Blood glucose increased",
    "Medical device malfunction",
]

EXCEL_PATH = "/home/cap/vigilex/eval_candidates/evaluation_pools_meddra_pt_candidates.xlsx"

def query_pts(conn, candidates):
    placeholders = ",".join(["%s"] * len(candidates))
    sql = f"""
        SELECT DISTINCT pt_code, pt_name, soc_name, meddra_version
        FROM processed.meddra_terms
        WHERE pt_name ILIKE ANY(ARRAY[{placeholders}])
        ORDER BY pt_name;
    """
    with conn.cursor() as cur:
        cur.execute(sql, candidates)
        return cur.fetchall()

def add_validation_sheet(wb, rows, candidates):
    # Remove existing sheet if present
    if "MedDRA_Validation" in wb.sheetnames:
        del wb["MedDRA_Validation"]

    ws = wb.create_sheet("MedDRA_Validation")

    # Header
    headers = ["pt_name_queried", "pt_code", "pt_name_exact", "soc_name", "meddra_version", "status"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font

    # Build lookup
    found = {r[1].lower(): r for r in rows}

    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")

    for r, cand in enumerate(candidates, 2):
        match = found.get(cand.lower())
        if match:
            ws.cell(r, 1, cand)
            ws.cell(r, 2, match[0])   # pt_code
            ws.cell(r, 3, match[1])   # pt_name exact
            ws.cell(r, 4, match[2])   # soc_name
            ws.cell(r, 5, match[3])   # meddra_version
            ws.cell(r, 6, "VALIDATED")
            for c in range(1, 7):
                ws.cell(r, c).fill = green
        else:
            ws.cell(r, 1, cand)
            ws.cell(r, 6, "NOT_FOUND -- check spelling or use fuzzy search")
            for c in range(1, 7):
                ws.cell(r, c).fill = red

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 40
    return found

def fill_pt_codes(wb, found):
    """Fill meddra_pt_code column in MedDRA_PT_candidates sheet."""
    ws = wb["MedDRA_PT_candidates"]
    # Find column indices
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col_pt_name = headers.get("candidate_primary_meddra_pt")
    col_pt_code = headers.get("meddra_pt_code")
    if not col_pt_name or not col_pt_code:
        print("WARNING: Could not find required columns in MedDRA_PT_candidates sheet")
        return

    for r in range(2, ws.max_row + 1):
        pt_name = ws.cell(r, col_pt_name).value
        if pt_name and pt_name.lower() in found:
            ws.cell(r, col_pt_code, found[pt_name.lower()][0])

def main():
    print(f"Connecting to DB...")
    conn = psycopg2.connect(DB_URL)

    print(f"Querying {len(PT_CANDIDATES)} PT candidates...")
    rows = query_pts(conn, PT_CANDIDATES)
    print(f"Found {len(rows)} matches in MedDRA.")
    conn.close()

    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    found = add_validation_sheet(wb, rows, PT_CANDIDATES)
    fill_pt_codes(wb, found)

    wb.save(EXCEL_PATH)
    print(f"Done. Saved: {EXCEL_PATH}")
    print(f"Validated: {len(found)} / {len(PT_CANDIDATES)} PT candidates")
    not_found = [c for c in PT_CANDIDATES if c.lower() not in found]
    if not_found:
        print(f"NOT FOUND in DB: {not_found}")

if __name__ == "__main__":
    main()
