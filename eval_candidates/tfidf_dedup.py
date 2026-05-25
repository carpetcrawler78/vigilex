"""
tfidf_dedup.py
Reads pool_LZG.csv, pool_QFG.csv, pool_OYC.csv from v2/
Runs TF-IDF + cosine similarity to flag near-duplicates.
Output: eval_candidates_v2.xlsx with color-coded similarity scores.

Run: python3 tfidf_dedup.py
Requires: pip install scikit-learn openpyxl pandas
"""

import os
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import openpyxl
from openpyxl.styles import PatternFill, Font

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
V2_DIR   = os.path.join(BASE_DIR, "v2")
OUT_PATH = os.path.join(BASE_DIR, "eval_candidates_v2.xlsx")

DEVICES   = ["LZG", "QFG", "OYC"]

# Per-device thresholds -- OYC/pacemaker reports are structurally templated
# so we need a higher bar before calling something a "duplicate"
SIM_HIGH = {"LZG": 0.80, "QFG": 0.80, "OYC": 0.92}
SIM_MED  = {"LZG": 0.60, "QFG": 0.60, "OYC": 0.75}


def load_pools():
    frames = []
    for dev in DEVICES:
        path = os.path.join(V2_DIR, f"pool_{dev}.csv")
        if not os.path.exists(path):
            print(f"WARNING: {path} not found, skipping.")
            continue
        df = pd.read_csv(path)
        df["device"] = dev
        frames.append(df)
        print(f"{dev}: {len(df)} records loaded")
    return pd.concat(frames, ignore_index=True)


PAIRWISE_LIMIT = 2000  # skip full pairwise matrix above this size (RAM safety)

def run_tfidf(df):
    """
    Compute max cosine similarity per record within same device group.
    Skips full pairwise matrix for large pools (> PAIRWISE_LIMIT) to
    avoid OOM -- greedy_diverse_select handles diversity for those.
    """
    vec = TfidfVectorizer(
        ngram_range=(1, 2),
        min_df=1,
        max_features=5000,
        sublinear_tf=True,
    )

    df["max_sim_within_device"] = -1.0
    df["most_similar_to"]       = ""
    df["dedup_flag"]            = "keep"

    for dev in df["device"].unique():
        mask  = df["device"] == dev
        texts = df.loc[mask, "mdr_text"].fillna("").tolist()
        idxs  = df.index[mask].tolist()

        if len(texts) < 2:
            continue

        if len(texts) > PAIRWISE_LIMIT:
            print(f"  {dev}: {len(texts)} records -- skipping pairwise matrix "
                  f"(> {PAIRWISE_LIMIT}), greedy selection handles diversity.")
            for global_idx in idxs:
                df.at[global_idx, "dedup_flag"] = "skipped_large_pool"
            continue

        tfidf_matrix = vec.fit_transform(texts)
        sim_matrix   = cosine_similarity(tfidf_matrix)
        np.fill_diagonal(sim_matrix, 0)

        for i, global_idx in enumerate(idxs):
            sims        = sim_matrix[i]
            max_sim     = float(sims.max())
            max_j       = int(sims.argmax())
            similar_key = df.loc[idxs[max_j], "mdr_report_key"]

            df.at[global_idx, "max_sim_within_device"] = round(max_sim, 3)
            df.at[global_idx, "most_similar_to"]       = similar_key

            high = SIM_HIGH[dev] if isinstance(SIM_HIGH, dict) else SIM_HIGH
            med  = SIM_MED[dev]  if isinstance(SIM_MED,  dict) else SIM_MED
            if max_sim >= high:
                df.at[global_idx, "dedup_flag"] = "LIKELY_DUPLICATE"
            elif max_sim >= med:
                df.at[global_idx, "dedup_flag"] = "SIMILAR_review"

    return df


def greedy_diverse_select(df, n_per_device=10):
    """
    For each device: iteratively pick the record that is
    least similar to already-selected records (MMR-style).
    Adds column 'greedy_selected' (True/False).
    Useful for OYC where almost everything is flagged as duplicate.
    """
    vec = TfidfVectorizer(ngram_range=(1, 2), min_df=1,
                          max_features=5000, sublinear_tf=True)
    df["greedy_selected"] = False
    df["greedy_rank"]     = -1

    for dev in df["device"].unique():
        mask  = df["device"] == dev
        idxs  = df.index[mask].tolist()
        texts = df.loc[mask, "mdr_text"].fillna("").tolist()
        n     = min(n_per_device, len(texts))

        tfidf_matrix = vec.fit_transform(texts).toarray()

        selected_local = []   # local indices into texts/idxs
        remaining      = list(range(len(texts)))

        # Start with record that has lowest avg similarity to all others
        sim_all = cosine_similarity(tfidf_matrix)
        np.fill_diagonal(sim_all, 0)
        avg_sim = sim_all.mean(axis=1)
        first   = int(np.argmin(avg_sim))
        selected_local.append(first)
        remaining.remove(first)

        for rank in range(1, n):
            if not remaining:
                break
            # For each candidate: compute max similarity to any selected
            sel_vecs = tfidf_matrix[selected_local]
            cand_vecs = tfidf_matrix[remaining]
            sim_to_selected = cosine_similarity(cand_vecs, sel_vecs).max(axis=1)
            best_local = remaining[int(np.argmin(sim_to_selected))]
            selected_local.append(best_local)
            remaining.remove(best_local)

        for rank, local_i in enumerate(selected_local):
            global_idx = idxs[local_i]
            df.at[global_idx, "greedy_selected"] = True
            df.at[global_idx, "greedy_rank"]     = rank

        print(f"  {dev}: {len(selected_local)} diverse records selected")

    return df


def write_excel(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eval_Candidates_v2"

    # Colors
    red    = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    green  = PatternFill("solid", fgColor="C6EFCE")
    blue   = PatternFill("solid", fgColor="4472C4")
    hfont  = Font(color="FFFFFF", bold=True)

    cols = [
        "device", "mdr_report_key", "manufacturer_canonical",
        "greedy_rank", "greedy_selected",
        "max_sim_within_device", "most_similar_to", "dedup_flag", "mdr_text"
    ]
    for c, h in enumerate(cols, 1):
        cell = ws.cell(1, c, h)
        cell.fill = blue
        cell.font = hfont

    df_sorted = df.sort_values(["device", "greedy_rank", "max_sim_within_device"],
                               ascending=[True, True, False])

    purple = PatternFill("solid", fgColor="D9B3FF")

    for r, (_, row) in enumerate(df_sorted.iterrows(), 2):
        flag = row["dedup_flag"]
        if row.get("greedy_selected"):
            fill = purple   # greedy-selected: top pick regardless of sim score
        elif flag == "LIKELY_DUPLICATE":
            fill = red
        elif flag == "SIMILAR_review":
            fill = yellow
        else:
            fill = green

        for c, col in enumerate(cols, 1):
            cell = ws.cell(r, c, row.get(col, ""))
            cell.fill = fill

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 80

    # Freeze header
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(1, 1, "Device")
    ws2.cell(1, 2, "Total")
    ws2.cell(1, 3, "LIKELY_DUPLICATE")
    ws2.cell(1, 4, "SIMILAR_review")
    ws2.cell(1, 5, "keep")

    for r, dev in enumerate(DEVICES, 2):
        sub = df[df["device"] == dev]
        ws2.cell(r, 1, dev)
        ws2.cell(r, 2, len(sub))
        ws2.cell(r, 3, (sub["dedup_flag"] == "LIKELY_DUPLICATE").sum())
        ws2.cell(r, 4, (sub["dedup_flag"] == "SIMILAR_review").sum())
        ws2.cell(r, 5, (sub["dedup_flag"] == "keep").sum())

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")
    print("\nSummary:")
    print(df.groupby(["device", "dedup_flag"]).size().unstack(fill_value=0))


def main():
    df = load_pools()
    print(f"\nTotal records: {len(df)}")
    df = run_tfidf(df)
    print("\nRunning greedy diverse selection (10 per device)...")
    df = greedy_diverse_select(df, n_per_device=10)
    write_excel(df)
    print(f"\nRecords flagged LIKELY_DUPLICATE: {(df['dedup_flag']=='LIKELY_DUPLICATE').sum()}")
    print(f"Records flagged SIMILAR_review:   {(df['dedup_flag']=='SIMILAR_review').sum()}")
    print(f"Records clean (keep):             {(df['dedup_flag']=='keep').sum()}")


if __name__ == "__main__":
    main()
